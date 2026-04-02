"""
Generate Spinly QA test case workbook.
Run: python3 spinly_test_cases.py
Output: spinly_test_cases.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Colour palette ────────────────────────────────────────────────────────────
C_HEADER_BG   = "1F3864"   # dark navy  – sheet header row
C_HEADER_FG   = "FFFFFF"
C_SECTION_BG  = "2E75B6"   # blue       – section title rows
C_SECTION_FG  = "FFFFFF"
C_ALT         = "EBF3FB"   # light blue – alternating row tint
C_BORDER      = "B8CCE4"
C_PASS_BG     = "E2EFDA"   # green hint – Expected Result column
C_CREDS_BG    = "FFF2CC"   # amber      – credentials sheet

def hex_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border(color=C_BORDER):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def header_font(size=10, bold=True, color=C_HEADER_FG):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def body_font(size=10, bold=False, color="000000"):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def wrap_align(h="left", v="top"):
    return Alignment(horizontal=h, vertical=v, wrap_text=True)


# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 1 – README / Legend
# ═══════════════════════════════════════════════════════════════════════════════
ws_readme = wb.active
ws_readme.title = "README"
ws_readme.sheet_view.showGridLines = False

readme_rows = [
    ["Spinly QA & Security Test Plan", ""],
    ["", ""],
    ["Purpose", "Comprehensive test cases covering login, RBAC, POS, order lifecycle,\nworkflow transitions, inventory, loyalty/promo, and security edge cases."],
    ["Application", "Spinly – Mobile-first Laundry POS built on Frappe v17"],
    ["Generated", "2026-03-22"],
    ["", ""],
    ["Sheet", "Contents"],
    ["Credentials", "Mock test user accounts for each role"],
    ["RBAC Matrix", "Permission matrix across all DocTypes and roles"],
    ["AUTH", "Login flow & session test cases (AUTH-01 to AUTH-09)"],
    ["POS", "POS page access test cases (POS-01 to POS-04)"],
    ["Orders", "Order creation, submission, cancellation RBAC (ORD-01 to ORD-05)"],
    ["Job Cards", "Job Card workflow transitions RBAC (JC-01 to JC-05)"],
    ["Config & Inventory", "Settings access + inventory deduct/restore (CFG-01 to INV-04)"],
    ["Loyalty & Promo", "Loyalty account and promo campaign RBAC (LOY-01 to PRO-03)"],
    ["Security", "Input validation, auth bypass, edge cases (SEC-01 to SEC-07)"],
    ["", ""],
    ["Columns", ""],
    ["TC ID", "Unique test case identifier"],
    ["Category", "Functional area"],
    ["Description", "One-line summary of what is being tested"],
    ["Preconditions", "State required before the test begins"],
    ["Steps", "Numbered actions the tester performs"],
    ["Expected Result", "Observable outcome that marks the test as PASSED"],
    ["Priority", "Critical / High / Medium / Low"],
    ["Status", "Blank – tester fills in: PASS / FAIL / BLOCKED / SKIP"],
    ["Notes", "Tester observations, bug IDs, etc."],
]

ws_readme.column_dimensions["A"].width = 22
ws_readme.column_dimensions["B"].width = 78

for r_idx, (a, b) in enumerate(readme_rows, 1):
    ca = ws_readme.cell(r_idx, 1, a)
    cb = ws_readme.cell(r_idx, 2, b)
    cb.alignment = wrap_align()
    if r_idx == 1:
        ca.font = Font(name="Calibri", size=14, bold=True, color=C_HEADER_BG)
        ws_readme.row_dimensions[r_idx].height = 22
    elif a in ("Sheet", "Columns"):
        for c in (ca, cb):
            c.font = header_font(size=10, color=C_HEADER_BG)
            c.fill = hex_fill("D6E4F0")
            c.border = thin_border()
    elif a and b:
        ca.font = body_font(bold=True)
        cb.font = body_font()


# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 2 – Credentials
# ═══════════════════════════════════════════════════════════════════════════════
ws_creds = wb.create_sheet("Credentials")
ws_creds.sheet_view.showGridLines = False

creds_headers = ["#", "Role", "Full Name", "Username / Email", "Password", "Notes"]
creds_data = [
    [1, "System Manager", "Admin Operator",  "admin@spinly.test",        "Admin@Spinly1",   "Super admin — setup/teardown only"],
    [2, "Laundry Manager","Meera Nair",       "meera.manager@spinly.test","Manager@Spinly2", "Day-to-day operations lead"],
    [3, "Laundry Staff",  "Ravi Kumar",       "ravi.staff@spinly.test",   "Staff@Spinly3",   "Counter / POS operator"],
    [4, "Laundry Staff",  "Divya Patel",      "divya.staff@spinly.test",  "Staff@Spinly4",   "Second staff user for cross-user isolation"],
    [5, "No Role",        "Unknown User",     "ghost@spinly.test",        "Ghost@Spinly5",   "Must be rejected – no desk access"],
]

col_widths_creds = [5, 20, 18, 32, 20, 42]
for i, w in enumerate(col_widths_creds, 1):
    ws_creds.column_dimensions[get_column_letter(i)].width = w

# header
for col, h in enumerate(creds_headers, 1):
    cell = ws_creds.cell(1, col, h)
    cell.font = header_font()
    cell.fill = hex_fill(C_HEADER_BG)
    cell.alignment = wrap_align("center", "center")
    cell.border = thin_border()
ws_creds.row_dimensions[1].height = 18

for r, row in enumerate(creds_data, 2):
    bg = C_CREDS_BG if r % 2 == 0 else "FFFFFF"
    for c, val in enumerate(row, 1):
        cell = ws_creds.cell(r, c, val)
        cell.font = body_font()
        cell.fill = hex_fill(bg)
        cell.alignment = wrap_align("left" if c > 1 else "center", "center")
        cell.border = thin_border()
    ws_creds.row_dimensions[r].height = 16


# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 3 – RBAC Matrix
# ═══════════════════════════════════════════════════════════════════════════════
ws_rbac = wb.create_sheet("RBAC Matrix")
ws_rbac.sheet_view.showGridLines = False

rbac_headers = ["Resource / DocType", "System Manager", "Laundry Manager", "Laundry Staff"]
rbac_data = [
    ["Spinly Settings",        "R + W",              "R only",             "✗"],
    ["Laundry Order",          "R W C S Cancel Amend Delete", "R W C S Cancel Amend", "C + S (no write post-create)"],
    ["Laundry Job Card",       "R W C S Cancel Amend Delete", "R W C S Cancel Amend", "R W C S"],
    ["Laundry Machine",        "R W C Delete",       "R + W",              "R only"],
    ["Laundry Consumable",     "R W C Delete",       "R + W",              "✗"],
    ["Inventory Restock Log",  "R W C Delete",       "R W C",              "✗"],
    ["Loyalty Account",        "R W C Delete",       "R + W",              "R only"],
    ["Promo Campaign",         "R W C Delete",       "R W C",              "✗"],
    ["Scratch Card",           "R W C Delete",       "R W C",              "R only"],
    ["Spinly POS Page",        "✓",                  "✓",                  "✓"],
    ["Workflow: Job Card Lifecycle – Sorting→Washing→Drying→Ironing→Ready", "all transitions", "all transitions", "✓ (all 4 in-process)"],
    ["Workflow: Mark Delivered", "all transitions",  "all transitions",    "✓"],
    ["Workflow: Cancel Job Card", "✓",               "✓",                  "✗"],
]

col_widths_rbac = [36, 30, 30, 30]
for i, w in enumerate(col_widths_rbac, 1):
    ws_rbac.column_dimensions[get_column_letter(i)].width = w

for col, h in enumerate(rbac_headers, 1):
    cell = ws_rbac.cell(1, col, h)
    cell.font = header_font()
    cell.fill = hex_fill(C_HEADER_BG)
    cell.alignment = wrap_align("center", "center")
    cell.border = thin_border()
ws_rbac.row_dimensions[1].height = 18

GREEN  = "E2EFDA"
RED    = "FCE4D6"
YELLOW = "FFF2CC"

def rbac_bg(val):
    if val in ("✗",): return RED
    if val in ("✓", "R only"): return YELLOW
    if "Delete" in val or "Cancel" in val: return "F4CCCC"
    return GREEN

for r, row in enumerate(rbac_data, 2):
    for c, val in enumerate(row, 1):
        cell = ws_rbac.cell(r, c, val)
        if c == 1:
            cell.font = body_font(bold=True)
            cell.fill = hex_fill(C_ALT if r % 2 == 0 else "FFFFFF")
        else:
            cell.font = body_font()
            cell.fill = hex_fill(rbac_bg(val))
            cell.alignment = wrap_align("center", "center")
        cell.border = thin_border()
    ws_rbac.row_dimensions[r].height = 18


# ═══════════════════════════════════════════════════════════════════════════════
# Helper – build a test-case sheet
# ═══════════════════════════════════════════════════════════════════════════════
TC_HEADERS = ["TC ID", "Category", "Description", "Preconditions", "Steps",
              "Expected Result", "Priority", "Status", "Notes"]
TC_COL_WIDTHS = [10, 20, 38, 36, 52, 52, 12, 12, 28]

PRIORITY_COLORS = {
    "Critical": "C00000",
    "High":     "FF0000",
    "Medium":   "FF9900",
    "Low":      "70AD47",
}

def add_tc_sheet(wb, sheet_name, sections):
    """
    sections = list of (section_title, [tc_rows])
    tc_row   = (tc_id, category, description, preconditions, steps, expected, priority)
    """
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    for i, w in enumerate(TC_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # header row
    for col, h in enumerate(TC_HEADERS, 1):
        cell = ws.cell(1, col, h)
        cell.font = header_font()
        cell.fill = hex_fill(C_HEADER_BG)
        cell.alignment = wrap_align("center", "center")
        cell.border = thin_border()
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    row_num = 2
    alt = False

    for section_title, tc_rows in sections:
        # section banner
        ws.merge_cells(start_row=row_num, start_column=1,
                       end_row=row_num, end_column=len(TC_HEADERS))
        cell = ws.cell(row_num, 1, f"  {section_title}")
        cell.font = header_font(size=10, color=C_SECTION_FG)
        cell.fill = hex_fill(C_SECTION_BG)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row_num].height = 16
        row_num += 1

        for tc in tc_rows:
            tc_id, category, desc, prereq, steps, expected, priority = tc
            bg = C_ALT if alt else "FFFFFF"
            alt = not alt

            values = [tc_id, category, desc, prereq, steps, expected, priority, "", ""]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row_num, col, val)
                cell.font = body_font()
                cell.fill = hex_fill(bg)
                cell.alignment = wrap_align()
                cell.border = thin_border()
                if col == 7 and val in PRIORITY_COLORS:   # Priority
                    cell.font = Font(name="Calibri", size=10, bold=True,
                                     color=PRIORITY_COLORS[val])
            ws.row_dimensions[row_num].height = 60
            row_num += 1

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# Data – AUTH
# ═══════════════════════════════════════════════════════════════════════════════
auth_sections = [
  ("Login Flow", [
    ("AUTH-01","Login","Successful login — Staff",
     "Spinly running; Ravi account exists",
     "1. Navigate to /login\n2. Enter ravi.staff@spinly.test / Staff@Spinly3\n3. Click Login",
     "Redirected to /desk; session cookie set; no setup wizard redirect","Critical"),
    ("AUTH-02","Login","Successful login — Manager",
     "Meera account exists",
     "1. Navigate to /login\n2. Enter meera.manager@spinly.test / Manager@Spinly2\n3. Click Login",
     "Redirected to /desk; Laundry Manager sidebar visible","Critical"),
    ("AUTH-03","Login","Successful login — System Manager",
     "Admin account exists",
     "1. Navigate to /login\n2. Enter admin@spinly.test / Admin@Spinly1\n3. Click Login",
     "Redirected to /desk; full System Manager menu visible","Critical"),
    ("AUTH-04","Login","Wrong password rejected",
     "Ravi account exists",
     "1. Enter ravi.staff@spinly.test\n2. Enter wrong password 'WrongPass1'\n3. Click Login",
     "'Invalid credentials' error shown; no session created","Critical"),
    ("AUTH-05","Login","Non-existent user rejected",
     "ghost@spinly.test does not exist",
     "1. Enter ghost@spinly.test / Ghost@Spinly5\n2. Click Login",
     "Login rejected with error message; no desk access","High"),
    ("AUTH-06","Login","Empty credentials blocked",
     "None",
     "1. Navigate to /login\n2. Leave both fields blank\n3. Click Login",
     "Validation prevents submission; no server call or 401 returned","High"),
    ("AUTH-07","Security","SQL injection in email field",
     "None",
     "1. Enter ' OR 1=1 -- in email field\n2. Enter any password\n3. Submit form",
     "Login rejected; no error stack trace exposed to browser","Critical"),
    ("AUTH-08","Login","Session persists across tab close",
     "User is logged in",
     "1. Log in as Ravi\n2. Close the browser tab\n3. Reopen app URL",
     "Session restored without re-login (within session timeout)","Medium"),
    ("AUTH-09","Login","Logout clears session",
     "User is logged in",
     "1. Log in as Ravi\n2. Click logout via Frappe menu\n3. Navigate to /spinly-pos",
     "Redirected to /login; no desk content accessible","High"),
  ]),
]
add_tc_sheet(wb, "AUTH", auth_sections)


# ═══════════════════════════════════════════════════════════════════════════════
# Data – POS
# ═══════════════════════════════════════════════════════════════════════════════
pos_sections = [
  ("POS Page Access", [
    ("POS-01","POS","Staff can access POS page",
     "Ravi logged in",
     "1. Navigate to /spinly-pos",
     "POS page loads; keypad and customer panel visible","Critical"),
    ("POS-02","POS","Manager can access POS page",
     "Meera logged in",
     "1. Navigate to /spinly-pos",
     "POS page loads without errors","High"),
    ("POS-03","POS","Unauthenticated user blocked",
     "No active session",
     "1. Without logging in, navigate to /spinly-pos",
     "Redirected to /login; POS not rendered","Critical"),
    ("POS-04","POS","POS loads masters on open",
     "Ravi logged in; services and payment methods exist in fixtures",
     "1. Open POS page\n2. Monitor browser Network tab",
     "get_pos_masters API called; services and payment methods populated in UI dropdowns","High"),
    ("POS-05","POS","Phone keypad appends digits correctly",
     "POS open",
     "1. Tap digits 1, 2, 3 on the keypad",
     "#sp-phone-value displays '123'","Medium"),
    ("POS-06","POS","Backspace removes last digit",
     "POS open; phone value is '123'",
     "1. Tap backspace button once",
     "#sp-phone-value displays '12'","Medium"),
    ("POS-07","POS","Customer search fires after 300ms debounce",
     "POS open; customer with phone 9876543210 exists",
     "1. Type 9876543210 one digit at a time\n2. Observe API calls in Network tab",
     "Only one get_customer_by_phone call fires (after 300ms pause), not one per digit","High"),
  ]),
]
add_tc_sheet(wb, "POS", pos_sections)


# ═══════════════════════════════════════════════════════════════════════════════
# Data – Orders
# ═══════════════════════════════════════════════════════════════════════════════
order_sections = [
  ("Order Creation & Submission", [
    ("ORD-01","Orders","Staff can create and submit an order via POS",
     "Ravi logged in; customer CUST-00001 exists; active service exists",
     "1. Open POS\n2. Enter CUST-00001 phone\n3. Add at least one item\n4. Tap Submit Order",
     "Order created with docstatus=1; Job Card auto-created; consumable stock reduced","Critical"),
    ("ORD-02","Orders","Staff cannot cancel a submitted order",
     "A submitted order exists; Ravi logged in",
     "1. Open submitted order in Frappe desk\n2. Attempt to click Cancel",
     "Cancel button absent; or PermissionError thrown on attempt","Critical"),
    ("ORD-03","Orders","Manager can cancel a submitted order",
     "A submitted order exists; Meera logged in",
     "1. Open submitted order\n2. Click Cancel\n3. Confirm cancellation",
     "Order cancelled (docstatus=2); machine load released; consumable stock restored","Critical"),
    ("ORD-04","Orders","System Manager can delete a draft order",
     "A draft order exists; admin logged in",
     "1. Open draft order in desk\n2. Click Delete",
     "Order removed from database; no orphaned records","High"),
    ("ORD-05","Orders","Staff cannot edit order fields after submission",
     "Submitted order exists; Ravi logged in",
     "1. Open submitted order\n2. Attempt to edit 'special_instructions'",
     "Field not editable (Laundry Staff has write:0 on Laundry Order)","High"),
    ("ORD-06","Orders","Order requires at least one item",
     "Ravi logged in",
     "1. Open POS\n2. Enter customer\n3. Submit without adding items",
     "POS shows validation alert; order not submitted","High"),
    ("ORD-07","Orders","Order lot number auto-generated on save",
     "New order being created",
     "1. Save/submit new order\n2. Check lot_number field",
     "lot_number follows pattern LOT-.YYYY.-.##### and is non-empty","Medium"),
  ]),
  ("ETA & Machine Assignment", [
    ("ORD-08","ETA","Machine is assigned on order save",
     "At least one Idle machine with sufficient capacity exists",
     "1. Create new order with total_weight_kg < machine capacity\n2. Save (before submit)",
     "assigned_machine field is populated; expected_ready_date is set","High"),
    ("ORD-09","ETA","No machine assigned when all at capacity",
     "All machines are at 100% load",
     "1. Create order with weight that would exceed all machine capacities\n2. Save",
     "assigned_machine is blank; expected_ready_date falls back to earliest countdown_timer_end + processing_minutes","Medium"),
  ]),
]
add_tc_sheet(wb, "Orders", order_sections)


# ═══════════════════════════════════════════════════════════════════════════════
# Data – Job Cards
# ═══════════════════════════════════════════════════════════════════════════════
jc_sections = [
  ("Workflow Transitions", [
    ("JC-01","Job Cards","Staff advances Sorting → Washing",
     "Job Card in 'Sorting' state exists; Ravi logged in",
     "1. Open job card\n2. Click 'Start Washing' action",
     "workflow_state changes to 'Washing'","Critical"),
    ("JC-02","Job Cards","Staff advances Washing → Drying",
     "Job Card in 'Washing' state",
     "1. Open job card\n2. Click 'Start Drying'",
     "workflow_state = 'Drying'","Critical"),
    ("JC-03","Job Cards","Staff advances Drying → Ironing",
     "Job Card in 'Drying' state",
     "1. Open job card\n2. Click 'Start Ironing'",
     "workflow_state = 'Ironing'","Critical"),
    ("JC-04","Job Cards","Staff marks job card Ready",
     "Job Card in 'Ironing' state",
     "1. Open job card\n2. Click 'Mark Ready'",
     "workflow_state = 'Ready'","Critical"),
    ("JC-05","Job Cards","Staff marks job card Delivered",
     "Job Card in 'Ready' state; Ravi logged in",
     "1. Open job card\n2. Click 'Mark Delivered'",
     "workflow_state = 'Delivered'; docstatus = 1","Critical"),
    ("JC-06","Job Cards","Manager can cancel a submitted job card",
     "Submitted job card exists; Meera logged in",
     "1. Open submitted job card\n2. Click Cancel",
     "Job card cancelled (docstatus=2)","High"),
    ("JC-07","Job Cards","Staff cannot cancel a job card",
     "Submitted job card; Ravi logged in",
     "1. Open submitted job card\n2. Look for Cancel button",
     "Cancel button absent; PermissionError if attempted via URL","High"),
    ("JC-08","Job Cards","Job card inherits order's special instructions",
     "Order has special_instructions set",
     "1. Submit order with special instructions\n2. Open resulting job card",
     "special_instructions field on job card matches order (fetch_from)","Medium"),
    ("JC-09","Job Cards","Customer tier badge shown on job card",
     "CUST-00001 has tier=Gold in Loyalty Account",
     "1. Submit order for CUST-00001\n2. Open resulting job card",
     "customer_tier_badge = 'Gold'","Medium"),
  ]),
]
add_tc_sheet(wb, "Job Cards", jc_sections)


# ═══════════════════════════════════════════════════════════════════════════════
# Data – Config & Inventory
# ═══════════════════════════════════════════════════════════════════════════════
cfg_sections = [
  ("Spinly Settings Access", [
    ("CFG-01","Config","System Manager can write Spinly Settings",
     "Admin logged in",
     "1. Open Spinly Settings\n2. Change tax_rate_pct to 5\n3. Save",
     "Settings saved; tax_rate_pct = 5","Critical"),
    ("CFG-02","Config","Manager cannot write Spinly Settings",
     "Meera logged in",
     "1. Open Spinly Settings",
     "Page loads read-only; Save button absent or disabled","Critical"),
    ("CFG-03","Config","Staff has no access to Spinly Settings",
     "Ravi logged in",
     "1. Navigate to Spinly Settings directly",
     "PermissionError or page not shown in sidebar","High"),
  ]),
  ("Inventory Deduction & Restoration", [
    ("INV-01","Inventory","Manager can update consumable stock directly",
     "Meera logged in; a consumable record exists",
     "1. Open a Laundry Consumable record\n2. Edit current_stock\n3. Save",
     "current_stock updated successfully","High"),
    ("INV-02","Inventory","Staff has no access to consumable records",
     "Ravi logged in",
     "1. Navigate to Laundry Consumable list",
     "List hidden or PermissionError; no records accessible","Critical"),
    ("INV-03","Inventory","Stock deducted when order is submitted",
     "Note current_stock of each active consumable; order ready to submit",
     "1. Submit an order with total_weight_kg = W\n2. Check current_stock of each consumable",
     "Each consumable's current_stock reduced by (consumption_per_kg × W)","Critical"),
    ("INV-04","Inventory","Stock restored when order is cancelled",
     "Submitted order exists; note stock levels after INV-03",
     "1. Log in as Manager\n2. Cancel the order from INV-03\n3. Check consumable stock",
     "current_stock restored to pre-submit value for each consumable","Critical"),
    ("INV-05","Inventory","Manager can log a restock",
     "Meera logged in; consumable exists",
     "1. Create new Inventory Restock Log\n2. Set quantity_added = 10\n3. Save",
     "Consumable's current_stock increases by 10; stock_before and stock_after auto-set","High"),
    ("INV-06","Inventory","Staff cannot create restock log",
     "Ravi logged in",
     "1. Attempt to create Inventory Restock Log",
     "PermissionError; record not created","High"),
  ]),
]
add_tc_sheet(wb, "Config & Inventory", cfg_sections)


# ═══════════════════════════════════════════════════════════════════════════════
# Data – Loyalty & Promo
# ═══════════════════════════════════════════════════════════════════════════════
loy_sections = [
  ("Loyalty Accounts", [
    ("LOY-01","Loyalty","Staff can read loyalty account",
     "Ravi logged in",
     "1. Navigate to Loyalty Account list\n2. Open any record",
     "Records visible; all fields read-only; no Save button","High"),
    ("LOY-02","Loyalty","Staff cannot write loyalty account",
     "Ravi logged in",
     "1. Open a Loyalty Account\n2. Attempt to edit 'tier' field",
     "Field not editable; PermissionError if attempted via API","Critical"),
    ("LOY-03","Loyalty","Manager can edit loyalty account",
     "Meera logged in",
     "1. Open a Loyalty Account\n2. Change tier to 'Silver'\n3. Save",
     "Saved successfully; tier = 'Silver'","High"),
    ("LOY-04","Loyalty","Points earned after order submission",
     "CUST-00001 has a Loyalty Account; order submitted",
     "1. Note current_balance before test\n2. Submit order for CUST-00001\n3. Check Loyalty Account",
     "current_balance increased by points_per_kg × weight + points_per_currency_unit × net_amount","High"),
    ("LOY-05","Loyalty","Tier upgrade from Bronze to Silver",
     "Customer's total_points_earned just below silver threshold (default 500)",
     "1. Submit enough orders to cross 500 total lifetime points\n2. Check Loyalty Account tier",
     "tier changes from 'Bronze' to 'Silver'","Medium"),
  ]),
  ("Promo Campaigns", [
    ("PRO-01","Promo","Manager can create promo campaign",
     "Meera logged in",
     "1. Create new Promo Campaign\n2. Set type=Flash Sale, discount_pct=10, valid dates\n3. Save",
     "Campaign created with is_active=1","High"),
    ("PRO-02","Promo","Staff cannot access promo campaign list",
     "Ravi logged in",
     "1. Navigate to Promo Campaign list",
     "PermissionError or list not accessible","Critical"),
    ("PRO-03","Promo","POS applies active promo at preview",
     "Active Flash Sale promo exists covering today; Ravi logged in",
     "1. Add items to POS\n2. Observe preview_order response",
     "promo_discount_amount > 0; applied_promo is set to campaign name","High"),
    ("PRO-04","Promo","Expired promo is not applied",
     "Promo exists with end_date = yesterday",
     "1. Submit order eligible for the expired promo\n2. Check discount fields",
     "promo_discount_amount = 0; applied_promo is blank","High"),
    ("PRO-05","Promo","Higher-priority promo wins when multiple apply",
     "Two active promos: P1 priority=1 (10% off), P2 priority=5 (5% off)",
     "1. Submit order eligible for both promos\n2. Check applied_promo field",
     "applied_promo = P2 (priority=5 wins); discount reflects P2's rate","Medium"),
  ]),
]
add_tc_sheet(wb, "Loyalty & Promo", loy_sections)


# ═══════════════════════════════════════════════════════════════════════════════
# Data – Security
# ═══════════════════════════════════════════════════════════════════════════════
sec_sections = [
  ("API & Input Security", [
    ("SEC-01","Security","POS API rejects unauthenticated call",
     "No active session",
     "1. Without a session cookie, call /api/method/spinly.api.get_pos_masters via curl or browser",
     "HTTP 403 returned; no data exposed in response","Critical"),
    ("SEC-02","Security","create_customer rejects non-numeric phone",
     "Ravi logged in",
     "1. Call spinly.api.create_customer with phone='abc123xyz'",
     "API returns validation error; customer not created","Critical"),
    ("SEC-03","Security","create_customer rejects phone shorter than 7 digits",
     "Ravi logged in",
     "1. Call create_customer with phone='123'",
     "API returns validation error; customer not created","High"),
    ("SEC-04","Security","create_customer rejects empty name",
     "Ravi logged in",
     "1. Call create_customer with name='' and valid phone",
     "API returns 'Customer name is required'; customer not created","High"),
    ("SEC-05","Security","Loyalty redemption cannot exceed balance",
     "Customer has current_balance=100",
     "1. Submit order via POS redeeming 500 points for a customer with 100 balance",
     "Order rejected or redemption capped; no negative balance created","Critical"),
    ("SEC-06","Security","ETA uses only submitted (docstatus=1) job cards",
     "One draft job card on machine MAC-01; no submitted job cards",
     "1. Call preview_order for an order on MAC-01\n2. Check ETA calculation",
     "Draft job card excluded from queue minutes; ETA not inflated","High"),
    ("SEC-07","Security","Machine timer commit persists to DB",
     "Machine MAC-01 has countdown_timer_end set to 1 minute ago; status=Running",
     "1. Trigger clear_completed_timers scheduled job\n2. Immediately query DB for MAC-01",
     "Machine status='Idle', current_load_kg=0; changes committed (not lost)","High"),
    ("SEC-08","Security","CSS injection blocked in print format",
     "Alert tag exists with color_code containing CSS expression",
     "1. Set alert_tag color_code to 'red; background:url(x) onerror=alert(1)'\n2. Print order",
     "color_code is sanitised or escaped; no JS execution in print view","Critical"),
    ("SEC-09","Security","SQL injection via phone search blocked",
     "POS open",
     "1. Enter ' OR 1=1 -- in phone keypad (via direct API call)\n2. Observe get_customer_by_phone",
     "Frappe parameterised query returns empty result; no data leak; no 500 error","Critical"),
  ]),
]
add_tc_sheet(wb, "Security", sec_sections)


# ═══════════════════════════════════════════════════════════════════════════════
# Save
# ═══════════════════════════════════════════════════════════════════════════════
output_path = "/workspaces/Frappe/spinly_test_cases.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
